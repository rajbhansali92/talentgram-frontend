import csv
import io
import uuid
import logging
from datetime import datetime, timezone
from typing import List, Dict, Any, Tuple, Optional
import openpyxl
from pymongo import UpdateOne, InsertOne

from core import db
from services.import_schema import IMPORT_FIELDS
from services.import_transformers import (
    transform_name, transform_phone, transform_instagram,
    transform_height, transform_gender, transform_location,
    transform_list, transform_integer, transform_boolean, clean_placeholder
)
from services.import_validators import validate_row
from services.import_duplicates import check_duplicates

logger = logging.getLogger(__name__)

TRANSFORMERS = {
    "name": transform_name,
    "email": clean_placeholder,
    "phone": transform_phone,
    "alternate_contact_number": transform_phone,
    "age": transform_integer,
    "dob": clean_placeholder,
    "gender": transform_gender,
    "height": transform_height,
    "location": transform_location,
    "ethnicity": clean_placeholder,
    "instagram_handle": transform_instagram,
    "instagram_followers": clean_placeholder,
    "bio": clean_placeholder,
    "work_links": transform_list,
    "skills": transform_list,
    "tags": transform_list
}

def auto_detect_headers(headers: List[str]) -> Dict[str, Optional[str]]:
    """Intelligently map CSV headers to Mongo fields using aliases."""
    mapping = {}
    for field, cfg in IMPORT_FIELDS.items():
        aliases = cfg.get("aliases", [])
        matched = None
        for h in headers:
            h_clean = h.strip().lower()
            if h_clean == field.lower() or h_clean in aliases:
                matched = h
                break
        mapping[field] = matched
    return mapping

def parse_file(content: bytes, filename: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Parse CSV or XLSX file and return (headers, raw_rows)."""
    raw_rows = []
    headers = []
    
    if filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = wb.active
        # Read header row
        for cell in next(sheet.iter_rows(max_row=1)):
            headers.append(str(cell.value or "").strip())
            
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                raw_rows.append(row_dict)
    else:
        # Assume CSV
        # Handle string decoding safely
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
            
        reader = csv.reader(io.StringIO(text))
        headers = [h.strip() for h in next(reader)]
        for row in reader:
            if any(cell.strip() for cell in row):
                row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                raw_rows.append(row_dict)
                
    return headers, raw_rows

def apply_transforms(raw_row: Dict[str, Any], field_mapping: Dict[str, Optional[str]]) -> Dict[str, Any]:
    """Apply column mappings and transform raw row data into sanitized model format."""
    transformed = {}
    for mongo_field, sheet_col in field_mapping.items():
        cfg = IMPORT_FIELDS[mongo_field]
        val = None
        if sheet_col and sheet_col in raw_row:
            val = raw_row[sheet_col]
            
        transformer = TRANSFORMERS.get(mongo_field)
        if transformer and val is not None:
            transformed[mongo_field] = transformer(val)
        else:
            transformed[mongo_field] = cfg.get("default")
            
    # Default tags/skills to list if not set
    for field in ("location", "work_links", "skills", "tags"):
        if field not in transformed or transformed[field] is None:
            transformed[field] = []
            
    return transformed

async def validate_and_detect_duplicates(
    raw_rows: List[Dict[str, Any]], 
    field_mapping: Dict[str, Optional[str]]
) -> Dict[str, Any]:
    """Validate all rows and run duplicate detection."""
    valid_rows = []
    error_rows = []
    warning_rows = []
    duplicate_rows = []
    
    for idx, raw_row in enumerate(raw_rows):
        row_num = idx + 2 # row number is 1-indexed header + 1
        transformed = apply_transforms(raw_row, field_mapping)
        
        # Validation
        v_res = validate_row(transformed)
        status = v_res["status"]
        errors = v_res["errors"]
        warnings = v_res["warnings"]
        
        
        # Media Verification
        from services.import_validators import verify_row_media
        media_errors = await verify_row_media(transformed)
        if media_errors:
            status = "error"
            for i, me in enumerate(media_errors):
                errors[f"media_{i}"] = me
                
        row_meta = {
            "row_number": row_num,
            "data": transformed,
            "original": {k: str(v) for k, v in raw_row.items() if v is not None},
            "errors": errors,
            "warnings": warnings
        }
        
        if status == "error":
            error_rows.append(row_meta)
        else:
            # Check for duplicates in DB
            dup = await check_duplicates(transformed)
            if dup:
                row_meta["duplicate"] = dup
                duplicate_rows.append(row_meta)
            elif status == "warning":
                warning_rows.append(row_meta)
            else:
                valid_rows.append(row_meta)
                
    return {
        "valid": valid_rows,
        "errors": error_rows,
        "warnings": warning_rows,
        "duplicates": duplicate_rows
    }

async def execute_talent_import(
    import_id: str,
    records: List[Dict[str, Any]],
    dup_actions: Dict[str, Any],
    admin_id: str
) -> Dict[str, int]:
    """Process import operations, performing inserts/updates while tracking progress and recording snapshots."""
    from services.import_sessions import update_import_progress, record_import_snapshot
    from services.import_duplicates import merge_duplicate_profile

    session = await db.import_sessions.find_one({"_id": import_id})
    start_row = 0
    inserted_count = 0
    updated_count = 0
    if session:
        start_row = session.get("processed_rows", 0)
        inserted_count = session.get("successful_rows", 0)

    skipped_count = 0
    processed_count = 0
    
    failed_rows = []
    
    for row in records:
        processed_count += 1
        if processed_count <= start_row:
            continue
        data = row["data"]
        dup_info = row.get("duplicate")
        row_num = row.get("row_number", 0)
        
        # Base document fields
        doc = {
            **data,
            "id": data.get("id") or str(uuid.uuid4()),
            "import_id": import_id,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "status": "SUBMITTED",
            "approval_status": "approved",
            "created_by": admin_id,
            "notes": "",
            "social_links": {},
            "availability": "available"
        }
        
        # Apply dynamic auto-labeling rules (Mumbai, Female, Tall)
        from services.import_transformers import apply_auto_label_rules
        doc = await apply_auto_label_rules(doc)
        
        # Ensure media field is initialized
        if "media" not in doc:
            doc["media"] = []
            
        try:
            if dup_info:
                existing_id = dup_info["existing_talent_id"]
                action = dup_actions.get(existing_id, "skip")
                
                if action == "skip":
                    skipped_count += 1
                    continue
                elif action == "create":
                    doc["id"] = str(uuid.uuid4())
                    await db.talents.insert_one(doc)
                    await record_import_snapshot(import_id, doc["id"], None, doc)
                    inserted_count += 1
                else:
                    # Retrieve the existing document to record snapshot
                    existing_doc = await db.talents.find_one({"id": existing_id})
                    if existing_doc:
                        # Apply duplicate merge logic
                        merged_doc = merge_duplicate_profile(existing_doc, doc, action)
                        merged_doc["updated_at"] = datetime.now(timezone.utc).isoformat()
                        merged_doc["import_id"] = import_id
                        
                        await db.talents.replace_one({"id": existing_id}, merged_doc)
                        await record_import_snapshot(import_id, existing_id, existing_doc, merged_doc)
                        updated_count += 1
                    else:
                        # Fallback if somehow not found
                        await db.talents.insert_one(doc)
                        await record_import_snapshot(import_id, doc["id"], None, doc)
                        inserted_count += 1
            else:
                await db.talents.insert_one(doc)
                await record_import_snapshot(import_id, doc["id"], None, doc)
                inserted_count += 1
                
        except Exception as e:
            failed_rows.append({
                "row_number": row_num,
                "data": data,
                "error": str(e)
            })
            logger.error(f"Error importing row {row_num}: {e}")

        # Update progress in db every 50 rows
        if processed_count % 50 == 0:
            await update_import_progress(
                import_id, 
                processed_count, 
                inserted_count + updated_count,
                failed_rows,
                "importing"
            )
            failed_rows = []

    # Final progress sync
    status_str = "completed"
    if failed_rows or processed_count == 0:
        # If there are any failed rows we set status appropriately
        status_str = "completed"
        
    await update_import_progress(
        import_id, 
        processed_count, 
        inserted_count + updated_count,
        failed_rows,
        status_str
    )
        
    return {
        "imported": inserted_count,
        "updated": updated_count,
        "skipped": skipped_count,
        "failed": len(failed_rows)
    }

async def rollback_import_records(import_id: str) -> int:
    """Restores pre-import state snapshots for updated records and deletes inserted records."""
    session = await db.import_sessions.find_one({"_id": import_id})
    if not session:
        # Fallback to delete-only if session doesn't exist
        res = await db.talents.delete_many({"import_id": import_id})
        return res.deleted_count
        
    restored_count = 0
    deleted_count = 0
    
    for snap in session.get("snapshots", []):
        talent_id = snap.get("talent_id")
        before = snap.get("before")
        
        if before:
            # Restore previous state
            await db.talents.replace_one({"id": talent_id}, before)
            restored_count += 1
        else:
            # New record, delete it
            await db.talents.delete_one({"id": talent_id})
            deleted_count += 1
            
    # Mark session as rolled back
    await db.import_sessions.update_one(
        {"_id": import_id},
        {"$set": {"status": "rolled_back"}}
    )
    
    return restored_count + deleted_count

